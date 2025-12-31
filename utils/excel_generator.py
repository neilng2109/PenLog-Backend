"""Excel Export Generator for Penetration Tracking"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

def generate_penetration_excel(project, penetrations):
    """
    Generate an Excel workbook for penetration tracking
    
    Args:
        project: Project object
        penetrations: List of Penetration objects
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    
    # === SUMMARY SHEET ===
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Colors
    navy_fill = PatternFill(start_color="243b53", end_color="243b53", fill_type="solid")
    teal_fill = PatternFill(start_color="14b8a6", end_color="14b8a6", fill_type="solid")
    light_gray_fill = PatternFill(start_color="f0f4f8", end_color="f0f4f8", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=12)
    bold_font = Font(bold=True, size=11)
    
    # Title
    ws_summary['A1'] = 'PENETRATION LOG REPORT'
    ws_summary['A1'].font = Font(bold=True, size=16, color="243b53")
    ws_summary.merge_cells('A1:D1')
    
    # Project Info
    ws_summary['A3'] = 'Ship Name:'
    ws_summary['B3'] = project.ship_name
    ws_summary['A4'] = 'Project:'
    ws_summary['B4'] = project.name
    ws_summary['A5'] = 'Location:'
    ws_summary['B5'] = project.drydock_location
    ws_summary['A6'] = 'Start Date:'
    ws_summary['B6'] = project.start_date.strftime('%d %B %Y') if project.start_date else 'N/A'
    ws_summary['A7'] = 'Embarkation:'
    ws_summary['B7'] = project.embarkation_date.strftime('%d %B %Y') if project.embarkation_date else 'N/A'
    ws_summary['A8'] = 'Report Date:'
    ws_summary['B8'] = datetime.now().strftime('%d %B %Y %H:%M')
    
    # Make labels bold
    for row in range(3, 9):
        ws_summary[f'A{row}'].font = bold_font
    
    # Statistics
    ws_summary['A10'] = 'STATISTICS'
    ws_summary['A10'].font = Font(bold=True, size=14, color="243b53")
    ws_summary.merge_cells('A10:D10')
    
    total = len(penetrations)
    status_counts = {
        'not_started': 0,
        'open': 0,
        'closed': 0,
        'verified': 0
    }
    
    contractors = {}
    decks = set()
    
    for pen in penetrations:
        status_counts[pen.status] = status_counts.get(pen.status, 0) + 1
        if pen.contractor:
            contractors[pen.contractor.name] = contractors.get(pen.contractor.name, 0) + 1
        if pen.deck:
            decks.add(pen.deck)
    
    completion_rate = ((status_counts['closed'] + status_counts['verified']) / total * 100) if total > 0 else 0
    
    ws_summary['A12'] = 'Total Penetrations:'
    ws_summary['B12'] = total
    ws_summary['A13'] = 'Not Started:'
    ws_summary['B13'] = status_counts['not_started']
    ws_summary['A14'] = 'Open:'
    ws_summary['B14'] = status_counts['open']
    ws_summary['A15'] = 'Closed:'
    ws_summary['B15'] = status_counts['closed']
    ws_summary['A16'] = 'Verified:'
    ws_summary['B16'] = status_counts['verified']
    ws_summary['A17'] = 'Completion Rate:'
    ws_summary['B17'] = f"{completion_rate:.1f}%"
    ws_summary['A18'] = 'Number of Contractors:'
    ws_summary['B18'] = len(contractors)
    ws_summary['A19'] = 'Decks Covered:'
    ws_summary['B19'] = len(decks)
    
    for row in range(12, 20):
        ws_summary[f'A{row}'].font = bold_font
    
    # Column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30
    
    # === PENETRATIONS SHEET ===
    ws_pens = wb.create_sheet("Penetrations")
    
    # Headers
    headers = ['Pen ID', 'Deck', 'Fire Zone', 'Frame', 'Location', 'Type', 'Contractor', 
               'Status', 'Priority', 'Opened At', 'Completed At', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws_pens.cell(row=1, column=col)
        cell.value = header
        cell.font = white_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    sorted_pens = sorted(penetrations, key=lambda x: (x.contractor.name if x.contractor else '', x.pen_id))
    
    for row_idx, pen in enumerate(sorted_pens, 2):
        ws_pens.cell(row=row_idx, column=1, value=pen.pen_id or '')
        ws_pens.cell(row=row_idx, column=2, value=pen.deck or '')
        ws_pens.cell(row=row_idx, column=3, value=pen.fire_zone or '')
        ws_pens.cell(row=row_idx, column=4, value=pen.frame or '')
        ws_pens.cell(row=row_idx, column=5, value=pen.location or '')
        ws_pens.cell(row=row_idx, column=6, value=pen.pen_type or '')
        ws_pens.cell(row=row_idx, column=7, value=pen.contractor.name if pen.contractor else '')
        
        # Status with color coding
        status_display = {
            'not_started': 'Not Started',
            'open': 'Open',
            'closed': 'Closed',
            'verified': 'Verified'
        }.get(pen.status, pen.status)
        
        status_cell = ws_pens.cell(row=row_idx, column=8, value=status_display)
        
        if pen.status == 'verified':
            status_cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
            status_cell.font = Font(color="065f46", bold=True)
        elif pen.status == 'closed':
            status_cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
            status_cell.font = Font(color="1e40af", bold=True)
        elif pen.status == 'open':
            status_cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            status_cell.font = Font(color="991b1b", bold=True)
        
        ws_pens.cell(row=row_idx, column=9, value=pen.priority or '')
        ws_pens.cell(row=row_idx, column=10, value=pen.opened_at.strftime('%Y-%m-%d %H:%M') if pen.opened_at else '')
        ws_pens.cell(row=row_idx, column=11, value=pen.completed_at.strftime('%Y-%m-%d %H:%M') if pen.completed_at else '')
        ws_pens.cell(row=row_idx, column=12, value=pen.notes or '')
        
        # Alternating row colors
        if row_idx % 2 == 0:
            for col in range(1, 13):
                if col != 8:  # Don't override status color
                    ws_pens.cell(row=row_idx, column=col).fill = light_gray_fill
    
    # Column widths
    column_widths = [12, 8, 12, 8, 25, 15, 20, 12, 10, 18, 18, 30]
    for idx, width in enumerate(column_widths, 1):
        ws_pens.column_dimensions[get_column_letter(idx)].width = width
    
    # Freeze top row
    ws_pens.freeze_panes = 'A2'
    
    # === BY CONTRACTOR SHEET ===
    ws_contractors = wb.create_sheet("By Contractor")
    
    # Headers
    ws_contractors['A1'] = 'Contractor'
    ws_contractors['B1'] = 'Total Pens'
    ws_contractors['C1'] = 'Not Started'
    ws_contractors['D1'] = 'Open'
    ws_contractors['E1'] = 'Closed'
    ws_contractors['F1'] = 'Verified'
    ws_contractors['G1'] = 'Completion %'
    
    for col in range(1, 8):
        cell = ws_contractors.cell(row=1, column=col)
        cell.font = white_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    contractor_stats = {}
    for pen in penetrations:
        if pen.contractor:
            name = pen.contractor.name
            if name not in contractor_stats:
                contractor_stats[name] = {
                    'total': 0,
                    'not_started': 0,
                    'open': 0,
                    'closed': 0,
                    'verified': 0
                }
            contractor_stats[name]['total'] += 1
            contractor_stats[name][pen.status] = contractor_stats[name].get(pen.status, 0) + 1
    
    row_idx = 2
    for contractor_name in sorted(contractor_stats.keys()):
        stats = contractor_stats[contractor_name]
        completed = stats['closed'] + stats['verified']
        completion_pct = (completed / stats['total'] * 100) if stats['total'] > 0 else 0
        
        ws_contractors.cell(row=row_idx, column=1, value=contractor_name)
        ws_contractors.cell(row=row_idx, column=2, value=stats['total'])
        ws_contractors.cell(row=row_idx, column=3, value=stats.get('not_started', 0))
        ws_contractors.cell(row=row_idx, column=4, value=stats.get('open', 0))
        ws_contractors.cell(row=row_idx, column=5, value=stats.get('closed', 0))
        ws_contractors.cell(row=row_idx, column=6, value=stats.get('verified', 0))
        ws_contractors.cell(row=row_idx, column=7, value=f"{completion_pct:.1f}%")
        
        # Alternating rows
        if row_idx % 2 == 0:
            for col in range(1, 8):
                ws_contractors.cell(row=row_idx, column=col).fill = light_gray_fill
        
        row_idx += 1
    
    # Column widths
    for col in range(1, 8):
        ws_contractors.column_dimensions[get_column_letter(col)].width = 18
    
    ws_contractors.freeze_panes = 'A2'
    
    # === BY DECK SHEET ===
    ws_decks = wb.create_sheet("By Deck")
    
    # Headers
    ws_decks['A1'] = 'Deck'
    ws_decks['B1'] = 'Total Pens'
    ws_decks['C1'] = 'Not Started'
    ws_decks['D1'] = 'Open'
    ws_decks['E1'] = 'Closed'
    ws_decks['F1'] = 'Verified'
    ws_decks['G1'] = 'Completion %'
    
    for col in range(1, 8):
        cell = ws_decks.cell(row=1, column=col)
        cell.font = white_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data
    deck_stats = {}
    for pen in penetrations:
        if pen.deck:
            deck = pen.deck
            if deck not in deck_stats:
                deck_stats[deck] = {
                    'total': 0,
                    'not_started': 0,
                    'open': 0,
                    'closed': 0,
                    'verified': 0
                }
            deck_stats[deck]['total'] += 1
            deck_stats[deck][pen.status] = deck_stats[deck].get(pen.status, 0) + 1
    
    row_idx = 2
    for deck_name in sorted(deck_stats.keys()):
        stats = deck_stats[deck_name]
        completed = stats['closed'] + stats['verified']
        completion_pct = (completed / stats['total'] * 100) if stats['total'] > 0 else 0
        
        ws_decks.cell(row=row_idx, column=1, value=deck_name)
        ws_decks.cell(row=row_idx, column=2, value=stats['total'])
        ws_decks.cell(row=row_idx, column=3, value=stats.get('not_started', 0))
        ws_decks.cell(row=row_idx, column=4, value=stats.get('open', 0))
        ws_decks.cell(row=row_idx, column=5, value=stats.get('closed', 0))
        ws_decks.cell(row=row_idx, column=6, value=stats.get('verified', 0))
        ws_decks.cell(row=row_idx, column=7, value=f"{completion_pct:.1f}%")
        
        if row_idx % 2 == 0:
            for col in range(1, 8):
                ws_decks.cell(row=row_idx, column=col).fill = light_gray_fill
        
        row_idx += 1
    
    for col in range(1, 8):
        ws_decks.column_dimensions[get_column_letter(col)].width = 18
    
    ws_decks.freeze_panes = 'A2'
    
    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer