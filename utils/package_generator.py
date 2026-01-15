"""Complete Package Generator - Excel with Cloudinary Photo Links"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

def generate_complete_package(project, penetrations, upload_folder):
    """
    Generate complete package: Excel with Cloudinary photo links
    
    Args:
        project: Project object
        penetrations: List of Penetration objects
        upload_folder: Not used (kept for backward compatibility)
    
    Returns:
        BytesIO object containing the Excel file
    """
    
    # === CREATE EXCEL FILE ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Penetrations with Photos"
    
    # Colors
    navy_fill = PatternFill(start_color="243b53", end_color="243b53", fill_type="solid")
    light_gray_fill = PatternFill(start_color="f0f4f8", end_color="f0f4f8", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    link_font = Font(color="0563C1", underline="single", size=10)
    
    # Headers
    headers = ['Pen ID', 'Deck', 'Fire Zone', 'Location', 'Type', 'Contractor', 
               'Status', 'Opening Photo', 'Closing Photo', 'Photo Count']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = white_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sort penetrations
    sorted_pens = sorted(penetrations, key=lambda x: (x.contractor.name if x.contractor else '', x.pen_id))
    
    # Process each penetration
    for row_idx, pen in enumerate(sorted_pens, 2):
        
        # Basic pen info
        ws.cell(row=row_idx, column=1, value=pen.pen_id or '')
        ws.cell(row=row_idx, column=2, value=pen.deck or '')
        ws.cell(row=row_idx, column=3, value=pen.fire_zone or '')
        ws.cell(row=row_idx, column=4, value=pen.location or '')
        ws.cell(row=row_idx, column=5, value=pen.pen_type or '')
        ws.cell(row=row_idx, column=6, value=pen.contractor.name if pen.contractor else '')
        
        # Status with color
        status_display = {
            'not_started': 'Not Started',
            'open': 'Open',
            'closed': 'Closed',
            'verified': 'Verified'
        }.get(pen.status, pen.status)
        
        status_cell = ws.cell(row=row_idx, column=7, value=status_display)
        
        if pen.status == 'verified':
            status_cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
            status_cell.font = Font(color="065f46", bold=True)
        elif pen.status == 'closed':
            status_cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
            status_cell.font = Font(color="1e40af", bold=True)
        elif pen.status == 'open':
            status_cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
            status_cell.font = Font(color="991b1b", bold=True)
        
        # Get photos for this pen
        from models import Photo
        photos = Photo.query.filter_by(penetration_id=pen.id).order_by(Photo.uploaded_at).all()
        
        opening_photo_url = None
        closing_photo_url = None
        all_photo_urls = []
        
        # Collect photo URLs (no downloading needed!)
        for idx, photo in enumerate(photos):
            photo_type = photo.photo_type or 'general'
            
            # Store Cloudinary URL directly
            all_photo_urls.append({
                'url': photo.filepath,
                'type': photo_type,
                'filename': photo.filename
            })
            
            # Categorize as opening or closing
            if 'opening' in photo_type.lower() or (not closing_photo_url and idx == 0):
                opening_photo_url = photo.filepath
            elif 'closing' in photo_type.lower() or (not opening_photo_url and idx == 0):
                closing_photo_url = photo.filepath
            
            # Assign first two if not categorized
            if idx == 0 and not opening_photo_url:
                opening_photo_url = photo.filepath
            elif idx == 1 and not closing_photo_url:
                closing_photo_url = photo.filepath
        
        # Add hyperlinks to photos (Cloudinary URLs - open in browser)
        if opening_photo_url:
            opening_cell = ws.cell(row=row_idx, column=8)
            opening_cell.value = "View Opening"
            opening_cell.hyperlink = opening_photo_url  # Direct Cloudinary URL
            opening_cell.font = link_font
            opening_cell.alignment = Alignment(horizontal='center')
        else:
            ws.cell(row=row_idx, column=8, value="-")
            ws.cell(row=row_idx, column=8).alignment = Alignment(horizontal='center')
        
        if closing_photo_url:
            closing_cell = ws.cell(row=row_idx, column=9)
            closing_cell.value = "View Closing"
            closing_cell.hyperlink = closing_photo_url  # Direct Cloudinary URL
            closing_cell.font = link_font
            closing_cell.alignment = Alignment(horizontal='center')
        else:
            ws.cell(row=row_idx, column=9, value="-")
            ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center')
        
        # Photo count
        folder_cell = ws.cell(row=row_idx, column=10)
        photo_count = len(all_photo_urls)
        folder_cell.value = f"{photo_count} photos" if photo_count > 0 else "-"
        folder_cell.alignment = Alignment(horizontal='center')
        folder_cell.font = Font(size=9, italic=True, color="666666")
        
        # Alternating row colors
        if row_idx % 2 == 0:
            for col in range(1, 11):
                if col not in [7, 8, 9]:  # Don't override status/link colors
                    ws.cell(row=row_idx, column=col).fill = light_gray_fill
    
    # Column widths
    column_widths = [12, 8, 12, 25, 15, 20, 12, 15, 15, 12]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Instructions", 0)
    
    instructions_text = [
        ["PenLog Complete Package", ""],
        ["", ""],
        ["EXCEL FILE WITH PHOTO LINKS", ""],
        ["Photos stored securely in the cloud and accessible via links.", ""],
        ["", ""],
        ["CONTENTS:", ""],
        ["1. This Excel file with penetration data", ""],
        ["2. Clickable photo links (opens in browser)", ""],
        ["", ""],
        ["HOW TO USE:", ""],
        ["1. Open this Excel file", ""],
        ["2. Enable editing if prompted by Excel", ""],
        ["3. Click blue 'View Opening' or 'View Closing' links", ""],
        ["4. Photos open in your web browser", ""],
        ["5. Requires internet connection to view photos", ""],
        ["", ""],
        ["BENEFITS:", ""],
        ["- Instant download (no waiting for photo packaging)", ""],
        ["- Small file size", ""],
        ["- Photos always up-to-date", ""],
        ["- Works on any device with internet", ""],
        ["", ""],
        ["TROUBLESHOOTING:", ""],
        ["- If Excel blocks links: Click 'Enable Editing' at top", ""],
        ["- If links don't open: Check your internet connection", ""],
        ["- Photos stored securely on Cloudinary", ""],
        ["", ""],
        ["PROJECT INFORMATION:", ""],
        ["Ship:", project.ship_name],
        ["Drydock:", project.drydock_location],
        ["Generated:", datetime.now().strftime('%d %B %Y %H:%M UTC')],
    ]
    
    for row_idx, row_data in enumerate(instructions_text, 1):
        ws_instructions.cell(row=row_idx, column=1, value=row_data[0])
        ws_instructions.cell(row=row_idx, column=2, value=row_data[1])
        
        if row_idx == 1:
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=16, color="243b53")
        elif ":" in row_data[0] and row_data[0].isupper():
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    
    ws_instructions.column_dimensions['A'].width = 40
    ws_instructions.column_dimensions['B'].width = 40
    
    # Save Excel file to buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer